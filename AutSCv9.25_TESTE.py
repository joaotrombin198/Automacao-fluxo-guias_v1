import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import sys
import time
import threading
import openpyxl
from openpyxl.styles import PatternFill, Font

fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')       # Verde claro
fill_erro = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')    # Vermelho claro


def iniciar_cronometro():
    start = time.time()
    def mostrar_tempo():
        while True:
            time.sleep(30)
            elapsed = time.time() - start
            print(f"[TEMPO] Tempo de execução: {int(elapsed)} segundos")
    t = threading.Thread(target=mostrar_tempo, daemon=True)
    t.start()

iniciar_cronometro()



# Caminho da planilha
CAMINHO_PLANILHA = r"C:\\Users\\joao.trombin\\Downloads\\Guias auditadas produção - Ressonâncias.xlsx"

carteirinha_fixa = "01987198000034000"  # já com dígito verificador incluso

def salvar_planilha_formatada(df, caminho_saida):
    print(f"[INFO] Salvando planilha atualizada em {caminho_saida} com formatação...")
    df.to_excel(caminho_saida, index=False)
    wb = openpyxl.load_workbook(caminho_saida)
    ws = wb.active

    fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')       # Verde claro
    fill_erro = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')    # Vermelho claro

    font_arial_normal = Font(name='Arial', bold=False)
    font_arial_bold = Font(name='Arial', bold=True)

    # Formatar coluna STATUS
    status_col = None
    for cell in ws[1]:
        if cell.value and str(cell.value).upper() == "STATUS":
            status_col = cell.column
            break
    if not status_col:
        print("[WARN] Coluna STATUS não encontrada para aplicar formatação.")
    else:
        for row in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row, column=status_col)
            valor = (status_cell.value or "").lower()
            if "erro" in valor or "falha" in valor or "inválido" in valor or "excluído" in valor:
                status_cell.fill = fill_erro
                status_cell.font = font_arial_bold
            elif valor.strip() != "":
                status_cell.fill = fill_ok
                status_cell.font = font_arial_normal
            else:
                # Sem preenchimento e fonte normal
                status_cell.fill = PatternFill(fill_type=None)
                status_cell.font = font_arial_normal

    # Formatar coluna NR_SEQ_SEGURADO como texto (para manter zeros à esquerda)
    col_carteira = None
    for cell in ws[1]:
        if cell.value and str(cell.value).upper() == "NR_SEQ_SEGURADO":
            col_carteira = cell.column
            break
    if col_carteira:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_carteira)
            cell.number_format = '@'  # Formato texto no Excel
            # Também força fonte Arial normal para essa coluna
            cell.font = font_arial_normal
    else:
        print("[WARN] Coluna NR_SEQ_SEGURADO não encontrada para formatar texto.")

    # Opcional: definir fonte Arial normal para todas as células, exceto o cabeçalho
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.font.name != 'Arial' or cell.font.bold:
                cell.font = font_arial_normal

    wb.save(caminho_saida)
    print("[INFO] Planilha salva e formatada com sucesso.")

def ajustar_carteira(carteira):
    carteira = carteira.strip()
    # Se começar com 198 e tiver 15 dígitos, adiciona 0 na frente para ficar 16 dígitos
    if carteira.startswith("198") and len(carteira) == 15:
        carteira = "0" + carteira
    return carteira

def validar_carteirinha(carteira):
    # Valida carteira fixa ou carteira com 16 dígitos e início 0198
    carteira = carteira.strip()
    if carteira == carteirinha_fixa:
        return True
    if carteira.startswith("0198") and len(carteira) == 16:
        return True
    return False

def extrair_segmentos_carteira(carteira):
    # Carteira fixa tem 17 dígitos com dígito verificador no final
    if carteira == carteirinha_fixa:
        unimed = carteira[0:4]
        cartao = carteira[4:8]
        benef = carteira[8:14]
        depen = carteira[14:16]
        digito_verificador = carteira[16]
        digitos_para_tentar = [digito_verificador]  # já conhecido
    else:
        # Carteira com 16 dígitos (sem dígito verificador)
        unimed = carteira[0:4]    # 0198
        cartao = carteira[4:8]
        benef = carteira[8:14]
        depen = carteira[14:16]
        digitos_para_tentar = list(range(10))  # tentar todos 0-9 para descobrir
    return unimed, cartao, benef, depen, digitos_para_tentar

def etapa_3_preencher_unimed_e_contratado(navegador):
    try:
        print("[ETAPA 3] Aguardando campo 'Unimed Executora'...")
        unimed_exec = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.ID, "cd_unimed_executora"))
        )
        unimed_exec.clear()
        unimed_exec.send_keys("0198")
        print("[ETAPA 3] Campo 'Unimed Executora' preenchido com 0198.")
        time.sleep(0.5)
    except TimeoutException:
        print("[ERRO - ETAPA 3] Campo 'Unimed Executora' não encontrado.")
        return False

    try:
        print("[ETAPA 3] Clicando na lupa para abrir popup do Contratado Solicitante...")
        lupa_prest_solic = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.localizadorImgLink"))
        )
        lupa_prest_solic.click()
        print("[ETAPA 3] Popup de busca do contratado aberto.")
        time.sleep(0.8)
    except TimeoutException:
        print("[ERRO - ETAPA 3] Lupa do Contratado Solicitante não encontrada ou não clicável.")
        return False

    main_window = navegador.current_window_handle
    WebDriverWait(navegador, 10).until(EC.number_of_windows_to_be(2))
    for handle in navegador.window_handles:
        if handle != main_window:
            popup = handle
            break
    navegador.switch_to.window(popup)
    print("[ETAPA 3] Troca para janela popup realizada.")
    time.sleep(0.8)

    try:
        print("[ETAPA 3] Preenchendo código da Unimed na popup...")
        cd_unimed_prestador = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.ID, "s_CD_UNIMED_PRESTADOR"))
        )
        cd_unimed_prestador.clear()
        cd_unimed_prestador.send_keys("0198")
        time.sleep(0.3)

        print("[ETAPA 3] Preenchendo nome do prestador na popup...")
        nome_prestador = navegador.find_element(By.ID, "s_NM_COMPLETO")
        nome_prestador.clear()
        nome_prestador.send_keys("MICHEL FARACO")
        time.sleep(0.3)

        print("[ETAPA 3] Clicando em localizar na popup...")
        botao_localizar = navegador.find_element(By.NAME, "Button_DoSearch")
        botao_localizar.click()
        time.sleep(0.3)
    except TimeoutException:
        print("[ERRO - ETAPA 3] Campos ou botão da popup não encontrados.")
        return False

    try:
        print("[ETAPA 3] Aguardando tabela com Michel Faraco...")
        link_michel = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//a[contains(text(), "MICHEL FARACO")]'))
        )
        link_michel.click()
        print("[ETAPA 3] Michel Faraco selecionado.")
        time.sleep(0.8)
    except TimeoutException:
        print("[ERRO - ETAPA 3] Link do Michel Faraco não encontrado na tabela.")
        return False

    # Voltar para janela principal
    navegador.switch_to.window(main_window)
    print("[ETAPA 3] Voltou para janela principal após seleção do contratado.")
    time.sleep(0.5)

    # Garante que está no frame correto antes de clicar no avançar
    try:
        navegador.switch_to.default_content()
        WebDriverWait(navegador, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "corpo")))
    except TimeoutException:
        print("[ERRO - ETAPA 3] Frame 'corpo' não encontrado para clicar no avançar.")
        return False

    # Tenta clicar no botão avançar com múltiplas abordagens
    try:
        print("[ETAPA 3] Tentando clicar no botão 'Avançar' da etapa 3...")

        # Tenta pelo id
        try:
            btn_avancar = WebDriverWait(navegador, 5).until(
                EC.element_to_be_clickable((By.ID, "botao_avancar"))
            )
            btn_avancar.click()
            print("[ETAPA 3] Botão 'Avançar' clicado via ID com sucesso.")
        except TimeoutException:
            # Tenta pelo atributo value (texto do botão)
            try:
                btn_avancar = WebDriverWait(navegador, 5).until(
                    EC.element_to_be_clickable((By.XPATH, '//input[@type="submit" and contains(@value, "Avançar")]'))
                )
                btn_avancar.click()
                print("[ETAPA 3] Botão 'Avançar' clicado via XPATH com sucesso.")
            except TimeoutException:
                # Fallback: clicar via javascript
                print("[ETAPA 3] Tentando clicar no botão 'Avançar' via JavaScript...")
                btn = navegador.find_element(By.XPATH, '//input[@type="submit" and contains(@value, "Avançar")]')
                navegador.execute_script("arguments[0].click();", btn)
                print("[ETAPA 3] Botão 'Avançar' clicado via JavaScript.")
        time.sleep(0.5)
    except Exception as e:
        print(f"[ERRO - ETAPA 3] Falha ao clicar no botão 'Avançar': {e}")
        return False

    return True


def etapa_4_preencher_campos(navegador, nr_seq_requisicao, cd_procedimento, qt_solicitado):
    try:
        print("[ETAPA 4] Aguardando dropdown 'Atendimento a RN'...")
        select_atend_rn = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.NAME, "FG_ATENDIMENTO_RN"))
        )
        for option in select_atend_rn.find_elements(By.TAG_NAME, 'option'):
            if option.get_attribute("value") == "N":
                option.click()
                print("[ETAPA 4] Selecionado 'Não' para Atendimento a RN.")
                break

        print("[ETAPA 4] Preenchendo campo Observações com NR_SEQ_REQUISICAO...")
        obs = navegador.find_element(By.ID, "ds_obs")
        obs.clear()
        obs.send_keys(str(nr_seq_requisicao))

        print("[ETAPA 4] Preenchendo campo Código do procedimento...")
        cd_servico = navegador.find_element(By.ID, "cd_servico_1")
        cd_servico.clear()
        cd_servico.send_keys(str(cd_procedimento))
        cd_servico.send_keys("\t")  # Tab para ativar validação
        time.sleep(1)

        print("[ETAPA 4] Verificando quantidade solicitada...")
        qtd_element = navegador.find_element(By.ID, "nr_qtd_1")
        qtd_informada = qtd_element.get_attribute("value")
        if str(qt_solicitado) != qtd_informada:
            print(f"[ETAPA 4] Quantidade solicitada ({qt_solicitado}) NÃO confere com valor da tela ({qtd_informada}). Atualizando na tela...")
            qtd_element.clear()
            qtd_element.send_keys(str(qt_solicitado))
            qtd_element.send_keys("\t")  # para disparar qualquer validação
            time.sleep(1)
        else:
            print(f"[ETAPA 4] Quantidade solicitada ({qt_solicitado}) confere com valor da tela ({qtd_informada}).")
        return True

    except TimeoutException:
        print("[ERRO - ETAPA 4] Elementos para preencher a etapa 4 não encontrados.")
        return False

def clicar_menu_sadt(navegador):
    try:
        navegador.switch_to.default_content()
        WebDriverWait(navegador, 10).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "/html/body/div/div[1]/iframe"))
        )
        botao_sadt = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[5]/ul/li[2]/a/div/div[2]'))
        )
        botao_sadt.click()
        print("[INFO] Menu SADT acessado para resetar tela.")
        time.sleep(0.6)
        return True
    except Exception as e:
        print(f"[ERRO] Falha ao acessar menu SADT: {e}")
        return False

def finalizar_solicitacao_tratando_erros(navegador):
    try:
        print("[FINALIZAR] Clicando no botão 'Finalizar'...")
        botao_finalizar = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.ID, "Button_Finalizar"))
        )
        botao_finalizar.click()
        time.sleep(1)

        # Verifica se apareceu mensagem de erro
        try:
            explica_hidden = WebDriverWait(navegador, 5).until(
                EC.presence_of_element_located((By.NAME, "explica_bloq_hidden"))
            )
            texto_erro = explica_hidden.get_attribute("value")
            if "erros" in texto_erro.lower():
                print("[FINALIZAR] Erros detectados na solicitação.")
                radio_sim = navegador.find_element(By.XPATH, '//input[@name="forcar_solic" and @value="1"]')
                radio_sim.click()
                time.sleep(1)
                print("[FINALIZAR] Marcado 'Sim' para finalizar mesmo com erros, clicando novamente em 'Finalizar'...")
                botao_finalizar = WebDriverWait(navegador, 10).until(
                    EC.element_to_be_clickable((By.ID, "Button_Finalizar"))
                )
                botao_finalizar.click()
                time.sleep(1.5)
        except TimeoutException:
            print("[FINALIZAR] Nenhum erro detectado após clicar em finalizar.")

        # Verifica se mensagem "Guia em estudo" aparece
        try:
            WebDriverWait(navegador, 5).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//td[contains(text(), "Em estudo") or contains(text(), "em estudo")]')
                )
            )
            print("[FINALIZAR] Solicitação finalizada com sucesso (mensagem 'em estudo' detectada).")
            if not clicar_menu_sadt(navegador):
                print("[ERRO] Não conseguiu voltar ao menu SADT.")
                return "falha", False
            return "ok", True

        except TimeoutException:
            print("[FINALIZAR] Não encontrou mensagem 'em estudo'. Tentando voltar ao menu SADT mesmo assim...")
            if not clicar_menu_sadt(navegador):
                print("[ERRO] Não conseguiu voltar ao menu SADT.")
                return "falha", False
            return "erro_envio", True

    except Exception as e:
        print(f"[FINALIZAR] Erro ao tentar finalizar: {e}")
        return "falha", False





# --- Carrega planilha ---
try:
    df = pd.read_excel(CAMINHO_PLANILHA, dtype={"NR_SEQ_SEGURADO": str})
    # Ajusta a carteira para 15 dígitos (sem zero) inicialmente - não altera aqui
    df["NR_SEQ_SEGURADO"] = df["NR_SEQ_SEGURADO"].str.strip()
    print("[INFO] Planilha carregada com sucesso!")
except FileNotFoundError:
    print(f"[ERRO] Planilha não encontrada: {CAMINHO_PLANILHA}")
    sys.exit(1)
except Exception as e:
    print(f"[ERRO] Falha ao carregar planilha: {e}")
    sys.exit(1)

# --- Inicializa navegador ---
navegador = webdriver.Chrome()
navegador.maximize_window()
navegador.get("https://rda-hml.unimedsc.com.br/autsc2/Login.do")
print("[INFO] Navegador iniciado e página acessada.")

# --- Login ---
try:
    WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.ID, "ds_login"))).send_keys("admin198")
    navegador.find_element(By.ID, "passwordTemp").send_keys("Unimed198@")
    WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.ID, "Button_DoLogin"))).click()
    print("[INFO] Login realizado com sucesso!")
except Exception as e:
    print(f"[ERRO] Falha no login: {e}")
    navegador.quit()
    sys.exit(1)

# --- Acessa menu SADT ---
if not clicar_menu_sadt(navegador):
    navegador.quit()
    sys.exit(1)

index = 0
tentativas = {}

while index < len(df):
    row = df.iloc[index]
    tentativas.setdefault(index, 0)

    try:
        tentativas[index] += 1
        print(f"\n[PROCESSO] Linha {index + 1} - Tentativa {tentativas[index]}")

        carteira_bruta = str(row["NR_SEQ_SEGURADO"]).strip()
        carteira_bruta = ajustar_carteira(carteira_bruta)

        print(f"[PROCESSO] Linha {index + 1} - Carteira bruta ajustada: {carteira_bruta}")

        if not validar_carteirinha(carteira_bruta):
            print(f"[AVISO] Carteira com formato inválido: {carteira_bruta}")
            df.at[index, "STATUS"] = "Formato inválido"
            salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))
            index += 1
            continue

        unimed, cartao, benef, depen, digitos_para_tentar = extrair_segmentos_carteira(carteira_bruta)
        print(f"[INFO] Segmentos extraídos: Unimed={unimed}, Cartao={cartao}, Beneficiário={benef}, Dependência={depen}")

        # Troca para o frame principal
        navegador.switch_to.default_content()
        WebDriverWait(navegador, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "corpo")))

        # Preencher campos do beneficiário
        navegador.find_element(By.ID, "bnf_cd_unimed").clear()
        navegador.find_element(By.ID, "bnf_cd_unimed").send_keys(unimed)
        navegador.find_element(By.ID, "bnf_cd_cartao").clear()
        navegador.find_element(By.ID, "bnf_cd_cartao").send_keys(cartao)
        navegador.find_element(By.ID, "bnf_cd_benef").clear()
        navegador.find_element(By.ID, "bnf_cd_benef").send_keys(benef)
        navegador.find_element(By.ID, "bnf_cd_dependencia").clear()
        navegador.find_element(By.ID, "bnf_cd_dependencia").send_keys(depen)

        sucesso = False
        for digito in digitos_para_tentar:
            print(f"[INFO] Testando dígito verificador: {digito}")
            navegador.find_element(By.ID, "bnf_cd_digito_verificador").clear()
            navegador.find_element(By.ID, "bnf_cd_digito_verificador").send_keys(str(digito))
            time.sleep(0.6)
            try:
                WebDriverWait(navegador, 1).until(
                    EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "2º Passo") or contains(text(), "2° Passo")]'))
                )
                print(f"[SUCESSO] Dígito verificador correto encontrado: {digito}")
                sucesso = True
                break
            except TimeoutException:
                print(f"[FALHA] Dígito {digito} incorreto.")

        if not sucesso:
            print("[AVISO] Nenhum dígito verificador válido encontrado.")
            df.at[index, "STATUS"] = "Digito inválido"
            salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))
            index += 1
            continue

        # Verifica se há data de exclusão
        try:
            data_exclusao_element = WebDriverWait(navegador, 5).until(
                EC.presence_of_element_located((By.XPATH, '//td[contains(text(), "Data Exclusão:")]/following-sibling::td[1]'))
            )
            data_exclusao_texto = data_exclusao_element.text.strip().replace('\xa0', '')
            print(f"[INFO] Data de Exclusão encontrada: '{data_exclusao_texto}'")
            if data_exclusao_texto != "__/__/____":
                print("[AVISO] Beneficiário com data de exclusão ativa. Usando carteirinha fixa.")
                df.at[index, "STATUS"] = "Data exclusão - Usou carteirinha fixa"
                if not clicar_menu_sadt(navegador):
                    raise Exception("Falha ao acessar menu SADT para resetar")

                carteira_bruta = carteirinha_fixa
                unimed = carteira_bruta[0:4]
                cartao = carteira_bruta[4:8]
                benef = carteira_bruta[8:14]
                depen = carteira_bruta[14:16]

                navegador.switch_to.default_content()
                WebDriverWait(navegador, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "corpo")))

                navegador.find_element(By.ID, "bnf_cd_unimed").clear()
                navegador.find_element(By.ID, "bnf_cd_unimed").send_keys(unimed)
                navegador.find_element(By.ID, "bnf_cd_cartao").clear()
                navegador.find_element(By.ID, "bnf_cd_cartao").send_keys(cartao)
                navegador.find_element(By.ID, "bnf_cd_benef").clear()
                navegador.find_element(By.ID, "bnf_cd_benef").send_keys(benef)
                navegador.find_element(By.ID, "bnf_cd_dependencia").clear()
                navegador.find_element(By.ID, "bnf_cd_dependencia").send_keys(depen)

                sucesso = False
                for digito in range(10):
                    print(f"[INFO] Testando dígito verificador na fixa: {digito}")
                    navegador.find_element(By.ID, "bnf_cd_digito_verificador").clear()
                    navegador.find_element(By.ID, "bnf_cd_digito_verificador").send_keys(str(digito))
                    time.sleep(0.6)
                    try:
                        WebDriverWait(navegador, 1).until(
                            EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "2º Passo") or contains(text(), "2° Passo")]'))
                        )
                        print(f"[SUCESSO] Dígito verificador da fixa correto: {digito}")
                        sucesso = True
                        break
                    except TimeoutException:
                        print(f"[FALHA] Dígito {digito} incorreto para fixa.")
                if not sucesso:
                    print("[ERRO] Não conseguiu validar dígito verificador para carteira fixa.")
                    df.at[index, "STATUS"] = "Erro fixa"
                    salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))
                    index += 1
                    continue

                print("[ETAPA 3 FIXA] Clicando no botão 'Avançar' após validar carteira fixa...")
                try:
                    botao_avancar = WebDriverWait(navegador, 5).until(EC.element_to_be_clickable((By.ID, "Button_Search")))
                    botao_avancar.click()
                    print("[ETAPA 3 FIXA] Botão 'Avançar' clicado com sucesso para carteira fixa.")
                    time.sleep(0.6)
                except TimeoutException:
                    print("[ERRO - ETAPA 3 FIXA] Botão 'Avançar' não encontrado para carteira fixa.")
                    df.at[index, "STATUS"] = "Erro avançar fixa"
                    salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))
                    index += 1
                    continue

            else:
                print("[INFO] Beneficiário ativo, segue com carteira da planilha.")
                df.at[index, "STATUS"] = "Carteirinha validada"
                print("[ETAPA 3 PLANILHA] Tentando clicar no botão 'Avançar' via JavaScript após validar carteira da planilha...")
                try:
                    botao_avancar = WebDriverWait(navegador, 5).until(EC.presence_of_element_located((By.ID, "Button_Search")))
                    navegador.execute_script("arguments[0].click();", botao_avancar)
                    print("[ETAPA 3 PLANILHA] Botão 'Avançar' clicado via JavaScript com sucesso para carteira da planilha.")
                    time.sleep(0.6)
                except TimeoutException:
                    print("[ERRO - ETAPA 3 PLANILHA] Botão 'Avançar' não encontrado.")
                    df.at[index, "STATUS"] = "Erro avançar planilha"
                    salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))
                    index += 1
                    continue

        except TimeoutException:
            print("[AVISO] Data de exclusão não encontrada, assumindo beneficiário ativo.")
            df.at[index, "STATUS"] = "Carteirinha validada"

        if not etapa_3_preencher_unimed_e_contratado(navegador):
            df.at[index, "STATUS"] = "Erro etapa 3"
            salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))
            index += 1
            continue

        nr_seq_requisicao = row["NR_SEQ_REQUISICAO"]
        cd_procedimento = row["CD_PROCEDIMENTO"]
        qt_solicitado = row["QT_SOLICITADO"]

        if not etapa_4_preencher_campos(navegador, nr_seq_requisicao, cd_procedimento, qt_solicitado):
            df.at[index, "STATUS"] = "Qtde divergente ou erro etapa 4"
            salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))
            index += 1
            continue
        
        resultado_finalizacao, sucesso = finalizar_solicitacao_tratando_erros(navegador)

        if not sucesso:
            df.at[index, "STATUS"] = "Erro ao finalizar"

        elif resultado_finalizacao == "ok":
            df.at[index, "STATUS"] = "Enviado"

        elif resultado_finalizacao == "erro_envio":
            df.at[index, "STATUS"] = "Enviado com erro"

        else:
            df.at[index, "STATUS"] = "Enviado (status indefinido)"

        salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))

        if not clicar_menu_sadt(navegador):
            print("[ERRO] Não conseguiu voltar para o menu SADT depois de finalizar.")
            df.at[index, "STATUS"] = "Finalizou mas não voltou ao menu"
            salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))

        index += 1
        time.sleep(0.6)

    except Exception as e:
        erro_msg = f"[ERRO FATAL] Exceção na linha {index + 1}: {e}"
        print(erro_msg)
        df.at[index, "STATUS"] = erro_msg
        salvar_planilha_formatada(df, CAMINHO_PLANILHA.replace(".xlsx", "_atualizado.xlsx"))
        navegador.quit()
        sys.exit(1)


print("[INFO] Processo finalizado.")
navegador.quit()